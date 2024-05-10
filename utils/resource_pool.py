import yaml

from pathlib import Path
from openpyxl import load_workbook

from utils.log import logger_manager
from utils.singleton import singleton
from utils.root_path import DEFAULT_CONFIG_PATH

logger = logger_manager.logger


@singleton
class ResourcePool:
    """
    The class of resource pool, to deserialize and load the resource
    """

    def __init__(self, resource_file_name):
        self.resource_file_name = resource_file_name
        self.resource_file_base_path = DEFAULT_CONFIG_PATH
        self.GROUPS = None
        self.ADMIN = None
        self.USER = None
        self.resource = self.load()

    def load(self):
        # Load environment specific config file
        resource_file_path = self.resource_file_base_path / self.resource_file_name

        if Path(resource_file_path).is_file():
            logger.info(
                f"Local config found {resource_file_path}",
            )
            yconfig = self.load_resource_from_yaml(resource_file_path)
            self.GROUPS = yconfig["groups"]["enable"]
            self.ADMIN = yconfig["administrators"]["wxid"]
            self.USER = yconfig.get("user", {})
            return yconfig
        else:
            logger.info(
                f"Local config not found {resource_file_path}",
            )
            assert False

    @staticmethod
    def load_resource_from_yaml(file_path):
        logger.info(f"Load resource, file name is {file_path}")
        with open(file_path, "r", encoding="utf-8") as my_resource:
            temp_resource = yaml.safe_load(my_resource)
        return temp_resource

    @staticmethod
    def write_resource_to_yaml(data, file_path):
        logger.info(f"Write resource, file name is {file_path}")
        with open(file_path, "w", encoding="utf-8") as file:
            yaml.dump(data, file)

    def rewrite_reload(self):
        resource_file_path = self.resource_file_base_path / self.resource_file_name
        self.write_resource_to_yaml(self.resource, resource_file_path)
        self.resource = self.load()

    def read_excel(self, file_name="scene.xlsx"):
        file_path = self.resource_file_base_path / file_name
        workbook = load_workbook(file_path)
        # 选择要读取的工作表
        sheet = workbook["Sheet1"]
        columns_index = {}
        result = {}
        error = []
        # 获取数据
        all_data = list(sheet.iter_rows(values_only=True))
        for row_number in range(len(all_data)):
            if row_number == 0:
                # 获取表头及其index
                for i in range(len(all_data[row_number])):
                    columns_index[all_data[row_number][i]] = i
                continue
            if all([
                all_data[row_number][columns_index["编号"]],
                all_data[row_number][columns_index["name"]],
                all_data[row_number][columns_index["description"]]
            ]):
                result[int(all_data[row_number][columns_index["编号"]])] = {
                    "name": all_data[row_number][columns_index["name"]],
                    "description": all_data[row_number][columns_index["description"]],
                }
            else:
                error.append(row_number + 1)
        return result, error





